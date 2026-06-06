"""Predefined ingest profiles — zero-config upload per portfolio company."""

from __future__ import annotations

import io
import re
from calendar import monthrange
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

import openpyxl

from company_registry import company_by_id, match_company
from csv_analyzer import ColumnMapping

MONTHS_NL = {
    "Jan": 1, "Feb": 2, "Mrt": 3, "Apr": 4, "Mei": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Okt": 10, "Nov": 11, "Dec": 12,
}

CANONICAL_HEADERS = [
    "date", "gl_account", "debit", "credit", "amount",
    "description", "journal", "document_no", "invoice_no", "source_sheet",
]


@dataclass
class ProfileParseResult:
    headers: list[str]
    rows: list[dict[str, str]]
    sheet_label: str
    workbook_profile: dict | None
    parser: str


def _iso(d: date | datetime) -> str:
    if isinstance(d, datetime):
        return d.date().isoformat()
    return d.isoformat()


def _cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def get_ingest_profile(opco_id: str) -> dict | None:
    company = company_by_id(opco_id)
    if not company:
        return None
    return company.get("ingest_profile")


def profile_to_enhancement(profile: dict, company: dict) -> dict[str, Any]:
    """Static enhancement dict — replaces Anthropic for trained profiles."""
    mapping = profile.get("column_mapping", {})
    return {
        "detected_system": profile.get("detected_system") or company.get("source_system"),
        "system_confidence": profile.get("confidence", 1.0),
        "column_mapping": mapping,
        "gl_suggestions": profile.get("gl_suggestions", []),
        "notes": profile.get("summary", ""),
        "ai_briefing": {
            "summary": profile.get("summary", ""),
            "dataType": profile.get("data_type", "transactions"),
            "targetStore": profile.get("target_store"),
            "storeReason": profile.get("store_reason", ""),
            "recommendedOpco": company.get("opco_name"),
            "recommendedCity": company.get("city"),
            "dateRange": None,
            "qualityChecks": profile.get("quality_checks", []),
            "mergeRecommendation": "ready" if profile.get("auto_merge") else "review_required",
            "controllerQuestion": profile.get("controller_question")
            or f"Merge {company.get('opco_name')} data into central database?",
        },
    }


def can_auto_merge(
    profile: dict | None,
    company_match: dict | None,
    normalized_count: int,
    block_merge: bool,
) -> tuple[bool, str]:
    if not company_match:
        return False, "Unknown file — not matched to a portfolio company"
    if not profile or not profile.get("trained"):
        return False, "No trained profile for this company"
    if normalized_count == 0:
        return False, "No valid rows after normalization"
    if block_merge:
        return False, "All rows already in database"
    if not profile.get("auto_merge"):
        return False, "Profile requires manual review"
    return True, "Trained profile — auto-merge"


def parse_exact_multi_year(content: bytes, company: dict) -> ProfileParseResult:
    from xlsx_reader import parse_workbook, workbook_profile_dict

    result = parse_workbook(content)
    return ProfileParseResult(
        headers=result.headers,
        rows=result.rows,
        sheet_label=result.primary_sheet,
        workbook_profile=workbook_profile_dict(result),
        parser="exact_multi_year",
    )


def parse_gilde_pl(content: bytes, company: dict) -> ProfileParseResult:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    sheet_names: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = next(ws.iter_rows(max_row=1, values_only=True), None)
        if not header:
            continue
        months = [h for h in header[1:13] if h in MONTHS_NL]
        if not months:
            continue
        sheet_names.append(sheet_name)
        year_match = re.search(r"\d{4}", sheet_name)
        year = 2026 if "2026" in sheet_name else int(year_match.group()) if year_match else 2025

        for row in ws.iter_rows(min_row=2, max_row=80, values_only=True):
            if not row or not row[0] or not isinstance(row[0], str):
                continue
            label = row[0].strip()
            if not label or not label[0].isdigit():
                continue
            gl = label.split()[0]
            for col_idx, month_label in enumerate(header[1:13], start=1):
                if month_label not in MONTHS_NL:
                    continue
                val = row[col_idx] if col_idx < len(row) else None
                if not val or not isinstance(val, (int, float)) or val == 0:
                    continue
                month = MONTHS_NL[month_label]
                day = min(15, monthrange(year, month)[1])
                txn_date = date(year, month, day)
                amount = round(float(val), 2)
                rows.append({
                    "date": txn_date.isoformat(),
                    "gl_account": gl,
                    "debit": "" if amount >= 0 else str(abs(amount)),
                    "credit": str(amount) if amount >= 0 else "",
                    "amount": str(amount),
                    "description": label,
                    "journal": "Gilde P&L",
                    "document_no": "",
                    "invoice_no": "",
                    "source_sheet": sheet_name,
                })
    wb.close()

    return ProfileParseResult(
        headers=CANONICAL_HEADERS,
        rows=rows,
        sheet_label=f"gilde ({len(sheet_names)} sheets)",
        workbook_profile={
            "mergedSheets": sheet_names,
            "sheetCount": len(sheet_names),
            "totalMergedRows": len(rows),
            "parser": "gilde_pl",
        },
        parser="gilde_pl",
    )


def parse_yuki_fintransactions(content: bytes, company: dict) -> ProfileParseResult:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    ws = wb.active
    gl_code = ""
    header_idx: int | None = None
    scanned: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        scanned.append(row)
        if row and row[0] == "Grootboekrekening" and row[1]:
            gl_code = str(row[1]).split(" - ")[0].strip()
        if header_idx is None and row and str(row[0]).strip() == "Nr.":
            header_idx = len(scanned) - 1

    if header_idx is not None and gl_code:
        sheet_title = ws.title
        for row in scanned[header_idx + 1 :]:
            if not row or not isinstance(row[2], datetime):
                continue
            debet = float(row[5] or 0)
            credit = float(row[6] or 0)
            amount = round(credit - debet, 2)
            if amount == 0:
                continue
            dagboek = str(row[4] or "")
            rows.append({
                "date": _iso(row[2]),
                "gl_account": gl_code,
                "debit": str(debet) if debet else "",
                "credit": str(credit) if credit else "",
                "amount": str(amount),
                "description": dagboek,
                "journal": dagboek,
                "document_no": _cell(row[0]),
                "invoice_no": "",
                "source_sheet": sheet_title,
            })
    else:
        sheet_title = wb.sheetnames[0] if wb.sheetnames else "yuki"
    wb.close()

    return ProfileParseResult(
        headers=CANONICAL_HEADERS,
        rows=rows,
        sheet_label=sheet_title,
        workbook_profile={"totalMergedRows": len(rows), "parser": "yuki_fintransactions"},
        parser="yuki_fintransactions",
    )


def parse_exact_gb_export(content: bytes, company: dict) -> ProfileParseResult:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        gl = str(int(row[0])) if isinstance(row[0], float) else str(row[0]).strip()
        txn_date = row[2] if len(row) > 2 else None
        if not isinstance(txn_date, datetime):
            continue
        debet = float(row[5] or 0) if len(row) > 5 else 0
        credit = float(row[6] or 0) if len(row) > 6 else 0
        amount = round(credit - debet, 2)
        if amount == 0:
            continue
        desc = str(row[7] or "GB export") if len(row) > 7 else "GB export"
        rows.append({
            "date": _iso(txn_date),
            "gl_account": gl,
            "debit": str(debet) if debet else "",
            "credit": str(credit) if credit else "",
            "amount": str(amount),
            "description": desc,
            "journal": "Exact GB",
            "document_no": "",
            "invoice_no": "",
            "source_sheet": ws.title,
        })
    wb.close()

    return ProfileParseResult(
        headers=CANONICAL_HEADERS,
        rows=rows,
        sheet_label=ws.title,
        workbook_profile={"totalMergedRows": len(rows), "parser": "exact_gb_export"},
        parser="exact_gb_export",
    )


PARSERS = {
    "exact_multi_year": parse_exact_multi_year,
    "gilde_pl": parse_gilde_pl,
    "yuki_fintransactions": parse_yuki_fintransactions,
    "exact_gb_export": parse_exact_gb_export,
}

# Catalog for onboarding UI — maps parser id → default ingest profile fields
PARSER_CATALOG: dict[str, dict[str, Any]] = {
    "exact_multi_year": {
        "label": "Exact — multi-year journals",
        "description": "Year tabs (2023–2026) + invoice sheet. Datum/Debet/Credit.",
        "source_systems": ["Exact"],
        "format_name": "Exact multi-year journals + invoice tab",
        "data_type": "revenue",
        "target_store": "revenue",
        "store_reason": "Verkoop journals and invoices — billing/revenue store",
        "multi_sheet": True,
    },
    "exact_gb_export": {
        "label": "Exact — GB export",
        "description": "GL col A, date col C, debit/credit cols F/G (8xxx revenue).",
        "source_systems": ["Exact"],
        "format_name": "Exact GB export (GL · date · debit/credit)",
        "data_type": "revenue",
        "target_store": "revenue",
        "store_reason": "GB 8xxx Omzet lines from Exact export",
        "multi_sheet": False,
    },
    "gilde_pl": {
        "label": "Gilde — monthly P&L",
        "description": "Rows = GL lines, columns = Dutch months (Jan–Dec). Unpivots to dated rows.",
        "source_systems": ["Gilde"],
        "format_name": "Gilde monthly P&L (rows × Jan–Dec columns)",
        "data_type": "mixed",
        "target_store": "mixed",
        "store_reason": "Gilde P&L spans materials, subs, billing, overhead — split by GL prefix",
        "multi_sheet": True,
    },
    "yuki_fintransactions": {
        "label": "Yuki — FinTransactions",
        "description": "GL in sheet header (Grootboekrekening), Nr./Datum/Debet/Credit columns.",
        "source_systems": ["Yuki"],
        "format_name": "Yuki FinTransactions (Grootboekrekening header)",
        "data_type": "mixed",
        "target_store": "mixed",
        "store_reason": "Yuki journal spans GL 4/5/8/9 — split by account on merge",
        "multi_sheet": False,
    },
    "generic": {
        "label": "Generic Excel/CSV (AI-assisted)",
        "description": "No custom parser — multi-sheet reader + AI column mapping on first upload.",
        "source_systems": ["Exact", "Yuki", "Gilde", "Snelstart", "Unknown"],
        "format_name": "Generic export (AI mapping on upload)",
        "data_type": "mixed",
        "target_store": "mixed",
        "store_reason": "Routed by GL category after AI/heuristic mapping",
        "multi_sheet": False,
    },
}

DEFAULT_COLUMN_MAPPING = {
    "date": "date",
    "gl_account": "gl_account",
    "debit": "debit",
    "credit": "credit",
    "description": "description",
}


def list_parsers_public() -> list[dict]:
    return [
        {"id": pid, **meta}
        for pid, meta in PARSER_CATALOG.items()
    ]


def build_ingest_profile(
    parser_id: str,
    opco_name: str,
    city: str,
    source_system: str,
    *,
    format_name: str | None = None,
    target_store: str | None = None,
    summary: str | None = None,
) -> dict:
    meta = PARSER_CATALOG.get(parser_id, PARSER_CATALOG["generic"])
    trained = parser_id != "generic" and parser_id in PARSERS
    fmt = format_name or meta["format_name"]
    store = target_store or meta["target_store"]
    profile: dict[str, Any] = {
        "trained": trained,
        "auto_merge": False,
        "confidence": 1.0 if trained else 0.75,
        "parser": parser_id if trained else None,
        "format_name": fmt,
        "file_types": ["xlsx", "csv"],
        "data_type": meta["data_type"],
        "target_store": store,
        "store_reason": meta["store_reason"],
        "detected_system": source_system,
        "column_mapping": dict(DEFAULT_COLUMN_MAPPING),
        "auto_approve_gl": trained,
        "summary": summary or f"{opco_name} — {fmt}. Location fixed: {city}.",
        "quality_checks": [
            f"{'Trained' if trained else 'AI-assisted'} format — review on first upload",
            f"Location fixed: {city} (not read from file columns)",
            f"Target store: {store}",
        ],
        "controller_question": f"Merge {opco_name} data into central database?",
    }
    if meta.get("multi_sheet"):
        profile["multi_sheet"] = True
    return profile


def parse_with_profile(content: bytes, filename: str, company_match: dict) -> ProfileParseResult | None:
    profile = get_ingest_profile(company_match["opcoId"])
    if not profile:
        return None
    parser_name = profile.get("parser")
    parser_fn = PARSERS.get(parser_name or "")
    if not parser_fn:
        return None
    company = company_by_id(company_match["opcoId"]) or {}
    return parser_fn(content, company)


def auto_gl_approvals(gl_suggestions: list[dict], profile: dict) -> dict[str, str]:
    if not profile.get("auto_approve_gl"):
        return {}
    out: dict[str, str] = {}
    for sug in gl_suggestions:
        gl = sug.get("glAccount") or sug.get("gl_account")
        cat = sug.get("suggestedCategory") or sug.get("category")
        if gl and cat and cat != "unmapped":
            out[gl] = cat
    defaults = profile.get("default_gl_categories") or {}
    for gl, cat in defaults.items():
        out[gl] = cat
    return out
