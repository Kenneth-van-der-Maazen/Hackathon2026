"""Multi-sheet Excel parsing for Altis subcompany / acquisition workbooks."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl

from csv_analyzer import journal_to_gl, scan_column_dates

YEAR_SHEET = re.compile(r"^\d{4}$")
CANONICAL_HEADERS = [
    "date",
    "gl_account",
    "debit",
    "credit",
    "amount",
    "description",
    "journal",
    "document_no",
    "invoice_no",
    "source_sheet",
]


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _norm_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]", "", h.lower().strip())


def _header_map(headers: list[str]) -> dict[str, str]:
    return {_norm_header(h): h for h in headers if h}


def _find_col(hmap: dict[str, str], *patterns: str) -> str | None:
    for pat in patterns:
        norm = _norm_header(pat)
        if norm in hmap:
            return hmap[norm]
        for key, orig in hmap.items():
            if norm in key or key in norm:
                return orig
    return None


@dataclass
class SheetProfile:
    name: str
    kind: str
    headers: list[str]
    row_count: int
    date_range: dict[str, str | None] = field(default_factory=lambda: {"start": None, "end": None})
    year_breakdown: dict[str, int] = field(default_factory=dict)
    opco_hint: str | None = None
    sample_rows: list[dict[str, str]] = field(default_factory=list)
    skipped_reason: str | None = None


@dataclass
class WorkbookParseResult:
    headers: list[str]
    rows: list[dict[str, str]]
    sheets: list[SheetProfile]
    merged_from: list[str]
    opco_hint: str | None = None
    city_hint: str | None = None
    primary_sheet: str = "merged"


def _try_yuki_fintransactions(ws) -> tuple[list[str], list[dict[str, str]], int] | None:
    gl_code = ""
    header_idx: int | None = None
    scanned: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        scanned.append(row)
        if row and row[0] == "Grootboekrekening" and row[1]:
            gl_code = str(row[1]).split(" - ")[0].strip()
        if header_idx is None and row and str(row[0]).strip() == "Nr.":
            header_idx = len(scanned) - 1
    if header_idx is None:
        return None

    header_row = scanned[header_idx]
    headers = [_cell_str(h) or f"col_{j}" for j, h in enumerate(header_row)]
    if "gl_account" not in headers:
        headers = [*headers, "gl_account"]

    rows: list[dict[str, str]] = []
    for row in scanned[header_idx + 1 :]:
        values = [_cell_str(v) for v in row]
        if not any(values):
            continue
        if values[0] == "" and len(values) > 2 and not values[2]:
            continue
        padded = values + [""] * max(0, len(headers) - 1 - len(values))
        record = dict(zip(headers[:-1], padded[: len(headers) - 1]))
        record["gl_account"] = gl_code
        rows.append(record)
    return headers, rows, len(rows) + 500


def _read_simple_sheet(ws, max_rows: int | None = None) -> tuple[list[str], list[dict[str, str]]]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [_cell_str(h) or f"col_{i}" for i, h in enumerate(header_row)]
    if not any(headers):
        return [], []

    rows: list[dict[str, str]] = []
    for i, row in enumerate(rows_iter):
        if max_rows is not None and i >= max_rows:
            break
        values = [_cell_str(v) for v in row]
        if not any(values):
            continue
        padded = values + [""] * max(0, len(headers) - len(values))
        rows.append(dict(zip(headers, padded[: len(headers)])))
    return headers, rows


def _extract_opco_hints(rows: list[list]) -> str | None:
    cities = ("Heeze", "Brunssum", "Andijk", "Winschoten", "Groningen", "Amsterdam")
    for row in rows:
        for cell in row:
            text = _cell_str(cell)
            if not text:
                continue
            for city in cities:
                if city.lower() in text.lower():
                    return city
            if "portfolio company" in text.lower():
                return text.strip()
            if text.strip() in cities:
                return text.strip()
    return None


def _classify_sheet(name: str, headers: list[str], rows: list[dict[str, str]]) -> str:
    hmap = {_norm_header(h) for h in headers}
    if YEAR_SHEET.match(name.strip()) and "datum" in hmap and ("debet" in hmap or "credit" in hmap):
        return "transaction"
    if "factuurdatum" in hmap or "factuurbedrag" in hmap:
        return "invoice"
    if "datum" in hmap and ("debet" in hmap or "credit" in hmap):
        return "transaction"
    if name.strip().lower().startswith("totaal") or name.strip().lower() == "summary":
        return "summary"
    if len(rows) >= 5 and any(_norm_header(h) in {"datum", "date", "factuurdatum"} for h in headers):
        return "transaction"
    if len(rows) < 2:
        return "skip"
    return "unknown"


def _to_canonical_transaction(row: dict[str, str], sheet_name: str) -> dict[str, str]:
    hmap = _header_map(list(row.keys()))
    date_col = _find_col(hmap, "datum", "date", "boekingsdatum")
    deb_col = _find_col(hmap, "debet", "debit")
    cred_col = _find_col(hmap, "credit")
    journal_col = _find_col(hmap, "dagboek", "journal")
    doc_col = _find_col(hmap, "bkst.nr.", "bkstnr", "document", "boekstuk")
    amt_col = _find_col(hmap, "bedrag", "amount", "saldo")

    journal = _cell_str(row.get(journal_col or "", ""))
    doc = _cell_str(row.get(doc_col or "", ""))
    desc_parts = [p for p in (journal, doc) if p]
    description = " · ".join(desc_parts) if desc_parts else "Imported transaction"

    return {
        "date": _cell_str(row.get(date_col or "", "")),
        "gl_account": journal_to_gl(journal) or "8000",
        "debit": _cell_str(row.get(deb_col or "", "")),
        "credit": _cell_str(row.get(cred_col or "", "")),
        "amount": _cell_str(row.get(amt_col or "", "")),
        "description": description,
        "journal": journal,
        "document_no": doc,
        "invoice_no": "",
        "source_sheet": sheet_name,
    }


def _to_canonical_invoice(row: dict[str, str], sheet_name: str) -> dict[str, str]:
    hmap = _header_map(list(row.keys()))
    date_col = _find_col(hmap, "factuurdatum", "datum", "date")
    inv_col = _find_col(hmap, "factuurnummer", "invoice", "factuur")
    amt_col = _find_col(hmap, "factuurbedrag", "bedrag", "amount", "credit")

    invoice_no = _cell_str(row.get(inv_col or "", ""))
    amount = _cell_str(row.get(amt_col or "", ""))

    return {
        "date": _cell_str(row.get(date_col or "", "")),
        "gl_account": "8000",
        "debit": "",
        "credit": amount,
        "amount": amount,
        "description": f"Invoice {invoice_no}" if invoice_no else "Sales invoice",
        "journal": "Verkoop",
        "document_no": invoice_no,
        "invoice_no": invoice_no,
        "source_sheet": sheet_name,
    }


def _profile_sheet(name: str, kind: str, headers: list[str], rows: list[dict[str, str]]) -> SheetProfile:
    date_col = None
    hmap = _header_map(headers)
    date_col = _find_col(hmap, "datum", "date", "factuurdatum", "boekingsdatum")

    year_breakdown: dict[str, int] = {}
    dates: list[str] = []
    if date_col:
        for row in rows:
            val = row.get(date_col, "")
            if not val:
                continue
            try:
                from unified_schema import parse_date

                iso = parse_date(val)
                dates.append(iso)
                year_breakdown[iso[:4]] = year_breakdown.get(iso[:4], 0) + 1
            except ValueError:
                continue
    elif rows:
        inferred, _ = scan_column_dates(headers, rows)
        if inferred:
            date_col = inferred
            for row in rows:
                val = row.get(inferred, "")
                if not val:
                    continue
                try:
                    from unified_schema import parse_date

                    iso = parse_date(val)
                    dates.append(iso)
                    year_breakdown[iso[:4]] = year_breakdown.get(iso[:4], 0) + 1
                except ValueError:
                    continue

    dates.sort()
    return SheetProfile(
        name=name,
        kind=kind,
        headers=headers,
        row_count=len(rows),
        date_range={"start": dates[0] if dates else None, "end": dates[-1] if dates else None},
        year_breakdown=year_breakdown,
        sample_rows=rows[:5],
    )


def parse_workbook(content: bytes) -> WorkbookParseResult:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    profiles: list[SheetProfile] = []
    canonical_rows: list[dict[str, str]] = []
    merged_from: list[str] = []
    opco_hint: str | None = None
    city_hint: str | None = None

    for name in wb.sheetnames:
        ws = wb[name]

        yuki = _try_yuki_fintransactions(ws)
        if yuki:
            headers, rows, _ = yuki
            kind = "yuki"
            for row in rows:
                canonical_rows.append({
                    "date": _find_col(_header_map(headers), "datum", "date") and row.get(
                        _find_col(_header_map(headers), "datum", "date") or "", ""
                    ) or "",
                    "gl_account": row.get("gl_account", ""),
                    "debit": row.get(_find_col(_header_map(headers), "debet", "debit") or "", ""),
                    "credit": row.get(_find_col(_header_map(headers), "credit") or "", ""),
                    "amount": "",
                    "description": row.get(_find_col(_header_map(headers), "omschrijving", "description") or "", "Yuki row"),
                    "journal": "",
                    "document_no": "",
                    "invoice_no": "",
                    "source_sheet": name,
                })
            profile = _profile_sheet(name, kind, headers, rows)
            profiles.append(profile)
            merged_from.append(name)
            continue

        headers, rows = _read_simple_sheet(ws)
        if not headers:
            profiles.append(SheetProfile(name=name, kind="skip", headers=[], row_count=0, skipped_reason="empty"))
            continue

        kind = _classify_sheet(name, headers, rows)

        if kind == "summary":
            scanned = list(ws.iter_rows(values_only=True, max_row=30))
            hint = _extract_opco_hints(scanned)
            if hint:
                city_hint = hint
                opco_hint = f"Portfolio Company {hint}" if "portfolio" not in hint.lower() else hint
            profiles.append(SheetProfile(name=name, kind=kind, headers=headers, row_count=len(rows), opco_hint=hint))
            continue

        if kind in ("transaction", "invoice"):
            profile = _profile_sheet(name, kind, headers, rows)
            profiles.append(profile)
            merged_from.append(name)
            for row in rows:
                if kind == "invoice":
                    canonical_rows.append(_to_canonical_invoice(row, name))
                else:
                    canonical_rows.append(_to_canonical_transaction(row, name))
            continue

        if kind == "skip":
            profiles.append(SheetProfile(name=name, kind=kind, headers=headers, row_count=len(rows), skipped_reason="too few rows"))
        else:
            profiles.append(SheetProfile(name=name, kind="unknown", headers=headers, row_count=len(rows), skipped_reason="unrecognized layout"))

    wb.close()

    # Drop rows without dates
    valid_rows = [r for r in canonical_rows if r.get("date")]

    return WorkbookParseResult(
        headers=CANONICAL_HEADERS,
        rows=valid_rows,
        sheets=profiles,
        merged_from=merged_from,
        opco_hint=opco_hint,
        city_hint=city_hint,
        primary_sheet=f"merged ({len(merged_from)} sheets)" if len(merged_from) > 1 else (merged_from[0] if merged_from else "none"),
    )


def xlsx_to_rows(content: bytes, max_rows: int | None = None) -> tuple[list[str], list[dict[str, str]], str]:
    """Parse workbook — merges all transaction/invoice year tabs."""
    result = parse_workbook(content)
    rows = result.rows if max_rows is None else result.rows[:max_rows]
    if not rows and not result.headers:
        raise ValueError("No usable worksheet found in Excel file")
    return result.headers, rows, result.primary_sheet


def workbook_profile_dict(result: WorkbookParseResult) -> dict:
    all_years: dict[str, int] = {}
    date_starts: list[str] = []
    date_ends: list[str] = []
    for sheet in result.sheets:
        for y, c in sheet.year_breakdown.items():
            all_years[y] = all_years.get(y, 0) + c
        if sheet.date_range.get("start"):
            date_starts.append(sheet.date_range["start"])
        if sheet.date_range.get("end"):
            date_ends.append(sheet.date_range["end"])

    return {
        "mergedSheets": result.merged_from,
        "sheetCount": len(result.sheets),
        "sheets": [
            {
                "name": s.name,
                "kind": s.kind,
                "rowCount": s.row_count,
                "headers": s.headers[:12],
                "dateRange": s.date_range,
                "yearBreakdown": s.year_breakdown,
                "opcoHint": s.opco_hint,
                "skippedReason": s.skipped_reason,
            }
            for s in result.sheets
        ],
        "totalMergedRows": len(result.rows),
        "yearBreakdown": all_years,
        "dateRange": {
            "start": min(date_starts) if date_starts else None,
            "end": max(date_ends) if date_ends else None,
        },
        "opcoHint": result.opco_hint,
        "cityHint": result.city_hint,
    }


def rows_to_csv_bytes(headers: list[str], rows: list[dict[str, str]]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def save_xlsx_as_csv(content: bytes, dest: Path) -> tuple[list[str], list[dict[str, str]], str, dict]:
    result = parse_workbook(content)
    dest.write_bytes(rows_to_csv_bytes(result.headers, result.rows))
    manifest = dest.parent / "workbook_manifest.json"
    manifest.write_text(
        __import__("json").dumps(workbook_profile_dict(result), indent=2),
        encoding="utf-8",
    )
    return result.headers, result.rows, result.primary_sheet, workbook_profile_dict(result)
